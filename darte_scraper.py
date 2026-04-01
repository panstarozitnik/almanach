# ============================================================
#  DARTESRO.SK Scraper – prihlásenie + vyhľadanie autora
#  Použitie: python darte_scraper.py
# ============================================================
#
#  INŠTALÁCIA (iba prvýkrát):
#  pip install requests beautifulsoup4 openpyxl
#
# ============================================================

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import time
from datetime import datetime

# ── Nastavenia ──────────────────────────────────────────────
DARTE_BASE    = "https://www.dartesro.sk"
DARTE_LOGIN   = "https://www.dartesro.sk/?act=login"
DARTE_BIO     = "https://www.dartesro.sk/?act=biografie"
DARTE_USER    = "jojourban"
DARTE_PASS    = "13491293428"

EXCEL_SUBOR   = "darte_databaza.xlsx"
HLAVICKY      = ["Meno autora", "Detail URL", "Popis / Bio úryvok", "Dátum pridania", "Hľadané slovo"]

FARBA_HLAVICKA = "1a1a2e"
FARBA_TEXT_HL  = "FFD700"
FARBA_RIADOK1  = "F8F4EE"
FARBA_RIADOK2  = "FFFFFF"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "sk-SK,sk;q=0.9,cs;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xhtml+xml;q=0.9,*/*;q=0.8",
    "Referer": DARTE_BASE,
}

# ── Session (udržiava cookies/prihlásenie) ───────────────────
session = requests.Session()
session.headers.update(HEADERS)


def prihlasit():
    """Prihlási sa na dartesro.sk a vráti True ak úspech."""
    print("🔐 Prihlasujem sa na dartesro.sk...")
    try:
        # Najprv načítame login stránku (získame prípadné CSRF tokeny)
        resp = session.get(DARTE_LOGIN, timeout=15)
        soup = BeautifulSoup(resp.text, "html.parser")

        # Nájdeme login formulár
        form = soup.find("form")
        payload = {}

        # Vyplníme všetky hidden polia (CSRF, atď.)
        if form:
            for inp in form.find_all("input"):
                name = inp.get("name", "")
                val  = inp.get("value", "")
                if name:
                    payload[name] = val

        # Nastavíme prihlasovacie údaje
        # Skúsime bežné názvy polí
        payload["username"] = DARTE_USER
        payload["password"] = DARTE_PASS
        payload["login"]    = DARTE_USER
        payload["pass"]     = DARTE_PASS
        payload["heslo"]    = DARTE_PASS
        payload["meno"]     = DARTE_USER

        # Zistíme action URL formulára
        action = form.get("action", DARTE_LOGIN) if form else DARTE_LOGIN
        if action.startswith("/"):
            action = DARTE_BASE + action
        elif not action.startswith("http"):
            action = DARTE_LOGIN

        resp2 = session.post(action, data=payload, timeout=15, allow_redirects=True)

        # Overíme prihlásenie – hľadáme logout link alebo meno užívateľa
        soup2 = BeautifulSoup(resp2.text, "html.parser")
        text = resp2.text.lower()

        if ("logout" in text or "odhlásiť" in text or "odhlasit" in text
                or DARTE_USER.lower() in text or "môj účet" in text):
            print("✅ Prihlásenie úspešné!")
            return True
        else:
            print("⚠️  Prihlásenie nepotvrdené – skúšam pokračovať bez overenia...")
            return True  # Skúsime aj tak

    except Exception as e:
        print(f"❌ Chyba pri prihlasovaní: {e}")
        return False


def hladaj_autora(slovo):
    """Vyhľadá autora v zozname biografií na dartesro.sk."""
    print(f"\n🔍 Hľadám autora '{slovo}' na dartesro.sk...")
    vysledky = []

    try:
        resp = session.get(DARTE_BIO, timeout=15)
        soup = BeautifulSoup(resp.text, "html.parser")

        # Hľadáme všetky linky/záznamy autorov
        # Skúsime rôzne selektory
        zaznamy = (
            soup.select(".biografia, .autor, .biografie, .author") or
            soup.select("table tr") or
            soup.select(".item, li a") or
            soup.find_all("a", href=lambda h: h and "biograf" in h.lower())
        )

        slovo_lower = slovo.lower()

        for zaznam in zaznamy:
            text = zaznam.get_text(strip=True)
            if not text or len(text) < 2:
                continue

            # Porovnáme či záznam obsahuje hľadané slovo
            if slovo_lower not in text.lower():
                continue

            # Nájdeme odkaz na detail
            link_el = zaznam if zaznam.name == "a" else zaznam.find("a")
            href = ""
            if link_el:
                href = link_el.get("href", "")
                if href.startswith("/"):
                    href = DARTE_BASE + href
                elif href and not href.startswith("http"):
                    href = DARTE_BASE + "/" + href

            # Načítame detail ak máme link
            popis = ""
            if href and href != DARTE_BASE:
                try:
                    time.sleep(0.5)
                    det = session.get(href, timeout=10)
                    det_soup = BeautifulSoup(det.text, "html.parser")
                    # Hľadáme hlavný textový obsah
                    content = (
                        det_soup.select_one(".content, .text, .bio, article, main, #content") or
                        det_soup.find("div", class_=lambda c: c and any(k in c.lower() for k in ["content","text","main","bio"]))
                    )
                    if content:
                        popis = content.get_text(separator=" ", strip=True)[:500]
                    else:
                        popis = det_soup.get_text(separator=" ", strip=True)[:300]
                except Exception:
                    popis = ""

            vysledky.append({
                "meno": text[:200],
                "href": href,
                "popis": popis,
            })

            print(f"  ✅ Nájdený: {text[:80]}")

    except Exception as e:
        print(f"❌ Chyba pri vyhľadávaní: {e}")

    return vysledky


# ── Excel ────────────────────────────────────────────────────
def nacitaj_excel():
    if os.path.exists(EXCEL_SUBOR):
        wb = openpyxl.load_workbook(EXCEL_SUBOR)
        ws = wb.active
        print(f"✅ Načítaný existujúci súbor: {EXCEL_SUBOR}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DARTE Autori"
        nastav_hlavicky(ws)
        print(f"✨ Vytvorený nový súbor: {EXCEL_SUBOR}")
    return wb, ws


def nastav_hlavicky(ws):
    for col, h in enumerate(HLAVICKY, 1):
        b = ws.cell(row=1, column=col, value=h)
        b.font = Font(bold=True, color=FARBA_TEXT_HL, size=11)
        b.fill = PatternFill("solid", fgColor=FARBA_HLAVICKA)
        b.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    sirky = [40, 60, 80, 20, 20]
    for i, s in enumerate(sirky, 1):
        ws.column_dimensions[get_column_letter(i)].width = s
    ws.row_dimensions[1].height = 30


def ziskaj_existujuce(ws):
    return {row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[1]}


def pridaj_do_excelu(ws, vysledky, slovo, existujuce):
    datum = datetime.now().strftime("%d.%m.%Y %H:%M")
    pridane = 0
    for v in vysledky:
        if v["href"] in existujuce:
            continue
        row_num = ws.max_row + 1
        data = [v["meno"], v["href"], v["popis"], datum, slovo]
        for col, val in enumerate(data, 1):
            b = ws.cell(row=row_num, column=col, value=val)
            farba = FARBA_RIADOK1 if row_num % 2 == 0 else FARBA_RIADOK2
            b.fill = PatternFill("solid", fgColor=farba)
            b.alignment = Alignment(wrap_text=True, vertical="top")
            b.font = Font(size=10)
            if col == 2 and val and val.startswith("http"):
                b.hyperlink = val
                b.font = Font(color="0563C1", underline="single", size=10)
        if v["href"]:
            existujuce.add(v["href"])
        pridane += 1
    return pridane


def formatuj(wb, ws):
    if ws.max_row < 2:
        return
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HLAVICKY))}{ws.max_row}"
    tenka = Side(style="thin", color="CCCCCC")
    from openpyxl.styles import Border
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(HLAVICKY)):
        for b in row:
            b.border = Border(left=tenka, right=tenka, top=tenka, bottom=tenka)


def main():
    print("=" * 55)
    print("   🖼️  DARTESRO.SK Scraper → Excel Databáza")
    print("=" * 55)

    # Prihlásenie
    ok = prihlasit()
    if not ok:
        print("❌ Prihlásenie zlyhalo. Ukončujem.")
        return

    wb, ws = nacitaj_excel()
    existujuce = ziskaj_existujuce(ws)
    print(f"📊 Záznamy v databáze: {len(existujuce)}")

    print("\nZadaj mená autorov (oddeľ čiarkou, napr: benka, bazovsky):")
    vstup = input("▶ ").strip()
    if not vstup:
        print("❌ Nezadal si žiadne slovo.")
        return

    slova = [s.strip() for s in vstup.split(",") if s.strip()]
    celkovo = 0

    for slovo in slova:
        vysledky = hladaj_autora(slovo)
        print(f"   📦 Nájdených: {len(vysledky)}")
        pridane = pridaj_do_excelu(ws, vysledky, slovo, existujuce)
        celkovo += pridane
        print(f"   ✅ Pridaných: {pridane}")

    formatuj(wb, ws)
    wb.save(EXCEL_SUBOR)
    print(f"\n💾 Uložené: {os.path.abspath(EXCEL_SUBOR)}")
    print(f"📊 Celkovo pridaných: {celkovo}")
    print(f"📊 Spolu v databáze: {ws.max_row - 1}")
    print("✅ Hotovo!")
    print("=" * 55)


if __name__ == "__main__":
    main()
